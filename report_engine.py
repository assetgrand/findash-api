"""
Backtest + raporty PDF/Excel (headless, pod FastAPI).
"""

from __future__ import annotations

import io
import tempfile
from datetime import datetime
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd

import core_analysis as core
import hybrid_engine as hybrid


def _sf(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        v = float(x)
        if v != v:
            return None
        return v
    except Exception:
        return None


def forecast_quality_backtest(ticker: str, horizon: str = "1M") -> Dict[str, Any]:
    """Walk-forward Hit% / MAE (ta sama logika co w tabelach desktopu)."""
    ticker = ticker.upper().strip()
    days = 63 if str(horizon).upper() in ("3M", "3", "90", "63") else 21
    horizon_label = "3M" if days > 30 else "1M"

    df = core.get_historical_prices(ticker, days=500)
    if df is None or getattr(df, "empty", True):
        raise ValueError(f"Brak danych dla {ticker}")
    df = core.calculate_indicators_on_df(df)

    sector = core.sector_mapping.get(ticker, "Default")
    fund_score = 50.0
    try:
        fa = core.get_comprehensive_fundamental_analysis(ticker)
        if fa and fa.get("combined_score") is not None:
            fund_score = float(fa["combined_score"])
    except Exception:
        pass

    q = core.backtest_forecast_quality(
        df, days_forward=days, sector=sector, fund_score=fund_score, ticker=ticker
    )
    if not q:
        raise ValueError("Za mało danych do wiarygodnego backtestu jakości prognoz")

    return {
        "ticker": ticker,
        "horizon": horizon_label,
        "days_forward": days,
        "sector": sector,
        "hit_rate": q.get("hit_rate"),
        "mae": q.get("mae"),
        "bias": q.get("bias"),
        "n_samples": q.get("n_samples"),
        "n_significant": q.get("n_significant"),
        "cover_rate": q.get("cover_rate"),
        "min_move": q.get("min_move"),
        "disclaimer": (
            "Hit% liczony tylko na istotnych ruchach (min_move). "
            "To ocena modelu, nie gwarancja przyszłych wyników."
        ),
    }


def strategy_backtest(ticker: str, initial_capital: float = 10000.0) -> Dict[str, Any]:
    """Prosty backtest MACD/RSI/ADX + SL/TP jak w desktopie."""
    ticker = ticker.upper().strip()
    df = core.get_historical_prices(ticker, days=500)
    if df is None or getattr(df, "empty", True):
        raise ValueError(f"Brak danych dla {ticker}")
    df = core.calculate_indicators_on_df(df)
    if df is None or len(df) < 30:
        raise ValueError("Za mało danych do backtestu strategii")

    df = df.copy()
    df["Signal"] = 0
    if "MACD" in df.columns and "MACD_Signal" in df.columns:
        df.loc[df["MACD"] > df["MACD_Signal"], "Signal"] = 1
        df.loc[df["MACD"] < df["MACD_Signal"], "Signal"] = -1
    if "RSI" in df.columns:
        df.loc[df["RSI"] < 30, "Signal"] = 1
        df.loc[df["RSI"] > 70, "Signal"] = -1
    if "ADX" in df.columns:
        df.loc[df["ADX"] < 20, "Signal"] = 0

    buy_hold_start = float(df["Close"].iloc[0])
    buy_hold_end = float(df["Close"].iloc[-1])
    buy_hold_return = ((buy_hold_end - buy_hold_start) / buy_hold_start) * 100.0

    strategy_capital = float(initial_capital)
    strategy_shares = 0.0
    strategy_trades: List[Dict[str, Any]] = []
    buy_price = 0.0

    for i in range(1, len(df)):
        current_price = float(df["Close"].iloc[i])
        signal = int(df["Signal"].iloc[i])
        dt = df.index[i]
        date_s = str(dt.date()) if hasattr(dt, "date") else str(dt)

        if signal == 1 and strategy_shares == 0:
            commission = current_price * 0.001
            strategy_shares = (strategy_capital - commission) / current_price
            strategy_capital = 0.0
            buy_price = current_price
            strategy_trades.append({"type": "BUY", "date": date_s, "price": round(current_price, 4)})
        elif signal == -1 and strategy_shares > 0:
            commission = current_price * 0.001
            strategy_capital = strategy_shares * current_price - commission
            strategy_shares = 0.0
            strategy_trades.append({"type": "SELL", "date": date_s, "price": round(current_price, 4)})
            buy_price = 0.0

        if strategy_shares > 0 and buy_price > 0:
            if current_price < buy_price * 0.95:
                commission = current_price * 0.001
                strategy_capital = strategy_shares * current_price - commission
                strategy_shares = 0.0
                strategy_trades.append({"type": "STOP LOSS", "date": date_s, "price": round(current_price, 4)})
                buy_price = 0.0
            elif current_price > buy_price * 1.10:
                commission = current_price * 0.001
                strategy_capital = strategy_shares * current_price - commission
                strategy_shares = 0.0
                strategy_trades.append({"type": "TAKE PROFIT", "date": date_s, "price": round(current_price, 4)})
                buy_price = 0.0

    if strategy_shares > 0:
        current_price = float(df["Close"].iloc[-1])
        commission = current_price * 0.001
        strategy_capital = strategy_shares * current_price - commission
        strategy_shares = 0.0
        dt = df.index[-1]
        date_s = str(dt.date()) if hasattr(dt, "date") else str(dt)
        strategy_trades.append({"type": "SELL (final)", "date": date_s, "price": round(current_price, 4)})

    strategy_return = ((strategy_capital - initial_capital) / initial_capital) * 100.0
    closed = [t for t in strategy_trades if t["type"] in ("SELL", "STOP LOSS", "TAKE PROFIT", "SELL (final)")]

    return {
        "ticker": ticker,
        "initial_capital": initial_capital,
        "final_value": round(strategy_capital, 2),
        "strategy_return_pct": round(strategy_return, 2),
        "buy_hold_return_pct": round(buy_hold_return, 2),
        "total_closed_trades": len(closed),
        "trades": strategy_trades[-40:],  # ostatnie 40, nie spam
        "disclaimer": "Backtest historyczny z prowizją 0.1%, SL 5%, TP 10%. Nie gwarantuje przyszłych wyników.",
    }


def build_report_payload(ticker: str) -> Dict[str, Any]:
    """Zbiera dane do PDF/Excel/JSON."""
    ticker = ticker.upper().strip()
    out: Dict[str, Any] = {"ticker": ticker, "generated_at": datetime.utcnow().isoformat() + "Z"}

    sector = core.sector_mapping.get(ticker, "Unknown")
    out["sector"] = sector

    try:
        a1 = None
        # reuse analyze path lightly
        df = core.get_historical_prices(ticker, days=500)
        if df is not None and not df.empty:
            df = core.calculate_indicators_on_df(df)
            last = df.iloc[-1]
            out["current_price"] = _sf(last["Close"])
            out["rsi"] = _sf(last["RSI"]) if "RSI" in df.columns else None
            out["macd"] = _sf(last["MACD"]) if "MACD" in df.columns else None
            out["adx"] = _sf(last["ADX"]) if "ADX" in df.columns else None
    except Exception as e:
        out["price_error"] = str(e)

    try:
        fa = core.get_comprehensive_fundamental_analysis(ticker)
        if fa:
            out["fundamental_rating"] = fa.get("fundamental_rating")
            out["combined_score"] = _sf(fa.get("combined_score"))
            out["company_score"] = _sf(fa.get("company_score"))
            out["country_score"] = _sf(fa.get("country_score"))
    except Exception as e:
        out["fundamental_error"] = str(e)

    for hz, key in (("1M", "forecast_1m"), ("3M", "forecast_3m")):
        try:
            days = 21 if hz == "1M" else 63
            df = core.get_historical_prices(ticker, days=500)
            df = core.calculate_indicators_on_df(df)
            fa = None
            try:
                fa = core.get_comprehensive_fundamental_analysis(ticker)
            except Exception:
                fa = {}
            try:
                pred, direction, chg = core.predict_with_technical_influence(
                    df, fa or {}, days, sector, ticker=ticker, quiet=True
                )
            except TypeError:
                pred, direction, chg = core.predict_with_technical_influence(df, fa or {}, days, sector)
            q = None
            try:
                q = core.backtest_forecast_quality(
                    df,
                    days_forward=days,
                    sector=sector,
                    fund_score=(fa or {}).get("combined_score") or 50,
                    ticker=ticker,
                )
            except Exception:
                pass
            out[key] = {
                "predicted_price": _sf(pred),
                "predicted_change_pct": _sf(chg),
                "direction": direction,
                "hit_rate": (q or {}).get("hit_rate"),
                "mae": (q or {}).get("mae"),
            }
        except Exception as e:
            out[key] = {"error": str(e)}

    try:
        out["hybrid"] = hybrid.analyze_hybrid(ticker, mode="Zrównoważony")
    except Exception as e:
        out["hybrid"] = {"error": str(e)}

    try:
        score = out.get("combined_score")
        out["perspective_3y"] = core.get_3year_perspective(ticker, score)
    except Exception as e:
        out["perspective_3y"] = {"error": str(e)}

    out["disclaimer"] = (
        "Raport automatyczny FinDash. Narzędzie analityczne – nie stanowi rekomendacji inwestycyjnej."
    )
    return out


def report_pdf_bytes(ticker: str) -> bytes:
    try:
        from fpdf import FPDF
    except ImportError as e:
        raise RuntimeError("Zainstaluj fpdf2") from e

    data = build_report_payload(ticker)
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    # Helvetica = bezpieczne na Linux (Render)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "FinDash - Raport analityczny", ln=True, align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Ticker: {ticker}", ln=True, align="C")
    pdf.cell(0, 6, f"Wygenerowano: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC", ln=True, align="C")
    pdf.ln(6)

    def line(label: str, value: Any):
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(55, 6, str(label)[:40], border=0)
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(0, 6, str(value) if value is not None else "-")

    line("Sektor", data.get("sector"))
    line("Cena", data.get("current_price"))
    line("RSI", data.get("rsi"))
    line("Rating", data.get("fundamental_rating"))
    line("Score", data.get("combined_score"))

    f1 = data.get("forecast_1m") or {}
    line("1M zmiana %", f1.get("predicted_change_pct"))
    line("1M kierunek", f1.get("direction"))
    line("1M Hit%", f1.get("hit_rate"))
    line("1M MAE", f1.get("mae"))

    f3 = data.get("forecast_3m") or {}
    line("3M zmiana %", f3.get("predicted_change_pct"))
    line("3M kierunek", f3.get("direction"))
    line("3M Hit%", f3.get("hit_rate"))

    hy = data.get("hybrid") or {}
    if "error" not in hy:
        line("Hybrid score", f"{hy.get('score')}/{hy.get('score_max')}")
        line("Hybrid sygnal", hy.get("signal"))
        line("Stop-loss", hy.get("stop_loss"))

    p3 = data.get("perspective_3y") or {}
    if isinstance(p3, dict) and "error" not in p3:
        line("Trend 3Y", p3.get("trend_3y"))
        line("Crash risk", p3.get("crash_risk_score"))
        line("Duza okazja", p3.get("duza_okazja"))
        line("Duze zagrozenie", p3.get("duze_zagrozenie"))

    pdf.ln(8)
    pdf.set_font("Helvetica", "I", 8)
    pdf.multi_cell(0, 5, data.get("disclaimer") or "")

    # fpdf output
    raw = pdf.output(dest="S")
    if isinstance(raw, str):
        return raw.encode("latin-1", errors="replace")
    return bytes(raw)


def report_excel_bytes(ticker: str) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("Zainstaluj openpyxl") from e

    data = build_report_payload(ticker)
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"
    ws.append(["Pole", "Wartosc"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    rows = [
        ("Ticker", data.get("ticker")),
        ("Sektor", data.get("sector")),
        ("Cena", data.get("current_price")),
        ("RSI", data.get("rsi")),
        ("Rating", data.get("fundamental_rating")),
        ("Combined score", data.get("combined_score")),
    ]
    f1 = data.get("forecast_1m") or {}
    rows += [
        ("1M %", f1.get("predicted_change_pct")),
        ("1M kierunek", f1.get("direction")),
        ("1M Hit%", f1.get("hit_rate")),
        ("1M MAE", f1.get("mae")),
    ]
    f3 = data.get("forecast_3m") or {}
    rows += [
        ("3M %", f3.get("predicted_change_pct")),
        ("3M kierunek", f3.get("direction")),
        ("3M Hit%", f3.get("hit_rate")),
    ]
    hy = data.get("hybrid") or {}
    if "error" not in hy:
        rows += [
            ("Hybrid score", hy.get("score")),
            ("Hybrid signal", hy.get("signal")),
            ("Stop loss", hy.get("stop_loss")),
        ]
    rows.append(("Disclaimer", data.get("disclaimer")))

    for a, b in rows:
        ws.append([a, b if not isinstance(b, (dict, list)) else str(b)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
